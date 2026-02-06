# -*- coding: utf-8 -*-
"""
email_sorter.py — Email Topic Organizer with Local BERT (Patched v2)

Key features
- Smart AI grouping with cosine confidence & auto-K (silhouette, with safe fallback)
- Learn from your Excel corrections (centroid persistence, per-Excel state path)
- TopicMap: “Merge Into” + “Custom Name” applied to learned centroids and final labels
- Multi-language (KO/EN/FR) using multilingual SentenceTransformer
- Clean, line-by-line Conversation Summary & Body Preview:
  Sender <email> | Monday, February 2, 2026 5:28 PM | key sentence (greetings removed)
- Message ID in Emails sheet (for robust learning & overrides)
- Safer exception handling (no bare `except:`), and neutral UI wording
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import os
import re
import hashlib
from datetime import datetime
from pathlib import Path
import email
from email import policy
from email.parser import BytesParser
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import numpy as np
import warnings
warnings.filterwarnings('ignore')


class EmailOrganizer:
    def __init__(self):
        self.emails = []
        self.excel_path = None  # GUI sets this when you choose an Excel file
        self.state_path = None  # <ExcelName>.learning_state.json (per-Excel)
        self.bert_model = None
        self.use_fallback_clustering = False
        self.correction_memory = {}              # by Message ID
        self.correction_memory_by_subject = {}   # by Subject
        self.topic_keywords = {}
        self.min_confidence = 0.65
        self.topicmap_rules = {}                 # {orig: {name: str, merge: str}}

    # ---------------------- Utilities: formatting & helpers ----------------------
    def _strip_greetings(self, text: str) -> str:
        """Remove typical greetings/boilerplate and return first meaningful sentence."""
        if not isinstance(text, str):
            return ""
        t = text.replace("\r\n", "\n").replace("\r", "\n")
        GREET_PREFIXES = (
            "안녕하세요", "안녕하십니까", "좋은 하루", "수고하십니다",
            "Hello", "Hi", "Dear", "Good morning", "Good afternoon", "Good evening",
            "Bonjour", "Bonsoir", "Salut",
        )
        JUNK_PREFIXES = (
            "주의", "CAUTION", "From:", "De :", "Sent:", "Envoyé", "Subject:", "Sujet:", "To:", "Cc:", "Date:"
        )
        lines = []
        for raw in t.split("\n"):
            ln = " ".join(raw.strip().split())
            if not ln:
                if lines and lines[-1] == "":
                    continue
                lines.append("")
                continue
            if any(ln.startswith(p) for p in JUNK_PREFIXES):
                continue
            low = ln.lower()
            if any(low.startswith(p.lower()) for p in GREET_PREFIXES):
                if len(ln) <= 20:
                    continue
                for p in GREET_PREFIXES:
                    if low.startswith(p.lower()):
                        ln = ln[len(p):].lstrip(" ,:—-")
                        break
            lines.append(ln)
        for ln in lines:
            if ln:
                return ln
        return ""

    def _format_sender_date_keyword(self, sender: str, date_str: str, key_sentence: str) -> str:
        s = (sender or "").strip()
        d = (date_str or "").strip()
        k = (key_sentence or "").strip()
        parts = [p for p in (s, d, k) if p]
        line = " | ".join(parts)
        return " ".join(line.split())[:1000]

    def _to_single_spaced_lines(self, text: str, max_lines: int = None, max_chars: int = None) -> str:
        if not isinstance(text, str):
            return ""
        t = text.replace("\r\n", "\n").replace("\r", "\n")
        out = []
        for raw in t.split("\n"):
            ln = " ".join(raw.strip().split())
            if ln == "":
                if out and out[-1] == "":
                    continue
            out.append(ln)
        if max_lines is not None and len(out) > max_lines:
            out = out[:max_lines]
        res = "\n".join(out)
        if max_chars is not None and len(res) > max_chars:
            res = res[:max_chars]
        return res

    def _state_path_for_excel(self, excel_path: str) -> Path:
        p = Path(excel_path)
        return p.with_name(f"{p.stem}.learning_state.json")

    # ---------------------- BERT init ----------------------
    def initialize_bert(self):
        try:
            import torch
            num_cores = os.cpu_count() or 4
            torch.set_num_threads(num_cores)
            from sentence_transformers import SentenceTransformer
            try:
                self.bert_model = SentenceTransformer('paraphrase-multilingual-MiniLM-L12-v2')
                return True
            except AttributeError as e:
                print(f"Primary model failed: {e}\nTrying alternative model...")
                try:
                    self.bert_model = SentenceTransformer('distiluse-base-multilingual-cased-v2')
                    return True
                except Exception as e2:
                    print(f"Alternative model failed: {e2}")
                    raise
        except ImportError:
            messagebox.showerror(
                "Missing Library",
                "sentence-transformers not found.\nInstall: pip install sentence-transformers torch"
            )
            return False
        except Exception as e:
            import traceback
            print("Error loading BERT:\n" + traceback.format_exc())
            response = messagebox.askyesno(
                "BERT Loading Error",
                "Could not load BERT (likely PyTorch issue).\n\n"
                "YES = Use improved K-Means (works now)\n"
                "NO  = Cancel and fix BERT installation"
            )
            if response:
                self.use_fallback_clustering = True
                return True
            return False

    def _ensure_bert(self):
        if not self.bert_model:
            if not self.initialize_bert():
                raise RuntimeError("BERT not available and fallback not allowed for learning.")

    # ---------------------- Parsing ----------------------
    def parse_email_file(self, filepath):
        try:
            if filepath.lower().endswith('.eml'):
                return self._parse_eml(filepath)
            elif filepath.lower().endswith('.msg'):
                return self._parse_msg(filepath)
            else:
                print(f"Unsupported file type: {filepath}")
                return None
        except Exception as e:
            print(f"Error parsing {filepath}: {e}")
            import traceback
            traceback.print_exc()
            return None

    def _parse_eml(self, filepath):
        with open(filepath, 'rb') as f:
            msg = BytesParser(policy=policy.default).parse(f)
        subject = msg.get('Subject', 'No Subject')
        from_addr = msg.get('From', 'Unknown')
        date_str = msg.get('Date', '')
        try:
            date_obj = email.utils.parsedate_to_datetime(date_str) if date_str else datetime.now()
        except Exception:
            date_obj = datetime.now()
        body = self._extract_body_eml(msg)
        topic_keyword = self._extract_topic_keyword(subject, body)
        thread_info = self._parse_conversation_thread(body)
        content_for_hash = f"{subject}{from_addr}{body[:500]}"
        msg_id = msg.get('Message-ID', hashlib.md5(content_for_hash.encode()).hexdigest())
        issue = self._extract_marker(subject + ' ' + body, r'(?i)(issue|bug|problem)\s*[:#]?\s*(\w+)')
        milestone = self._extract_marker(subject + ' ' + body, r'(?i)(milestone|phase|sprint)\s*[:#]?\s*(\w+)')
        pretty_date = date_obj.strftime("%A, %B %d, %Y %I:%M %p")
        key_sentence = self._strip_greetings(body if body else subject)
        body_preview_line = self._format_sender_date_keyword(from_addr, pretty_date, key_sentence)
        return {
            'message_id': msg_id,
            'subject': subject,
            'from': from_addr,
            'date': date_obj,
            'body': body,
            'issue': issue,
            'milestone': milestone,
            'filepath': filepath,
            'thread_count': thread_info['count'],
            'thread_summary': thread_info['summary'],
            'conversation_dates': thread_info['dates'],
            'topic_keyword': topic_keyword,
            'body_preview_line': body_preview_line,
        }

    def _parse_msg(self, filepath):
        try:
            import extract_msg
            msg = extract_msg.Message(filepath)
            subject = msg.subject or 'No Subject'
            from_addr = msg.sender or 'Unknown'
            try:
                date_obj = msg.date
                if date_obj is None:
                    date_obj = datetime.now()
                elif isinstance(date_obj, str):
                    try:
                        date_obj = datetime.strptime(date_obj, '%a, %d %b %Y %H:%M:%S %z')
                    except Exception:
                        try:
                            date_obj = datetime.strptime(date_obj, '%Y-%m-%d %H:%M:%S')
                        except Exception:
                            date_obj = datetime.now()
            except Exception:
                date_obj = datetime.now()
            body = msg.body or ''
            if not body:
                body = msg.htmlBody or ''
            topic_keyword = self._extract_topic_keyword(subject, body)
            thread_info = self._parse_conversation_thread(body)
            content_for_hash = f"{subject}{from_addr}{body[:500]}"
            msg_id = msg.messageId or hashlib.md5(content_for_hash.encode()).hexdigest()
            issue = self._extract_marker(subject + ' ' + body, r'(?i)(issue|bug|problem)\s*[:#]?\s*(\w+)')
            milestone = self._extract_marker(subject + ' ' + body, r'(?i)(milestone|phase|sprint)\s*[:#]?\s*(\w+)')
            pretty_date = date_obj.strftime("%A, %B %d, %Y %I:%M %p")
            key_sentence = self._strip_greetings(body if body else subject)
            body_preview_line = self._format_sender_date_keyword(from_addr, pretty_date, key_sentence)
            msg.close()
            return {
                'message_id': msg_id,
                'subject': subject,
                'from': from_addr,
                'date': date_obj,
                'body': body,
                'issue': issue,
                'milestone': milestone,
                'filepath': filepath,
                'thread_count': thread_info['count'],
                'thread_summary': thread_info['summary'],
                'conversation_dates': thread_info['dates'],
                'topic_keyword': topic_keyword,
                'body_preview_line': body_preview_line,
            }
        except ImportError:
            raise Exception("extract-msg library is required for .msg files.")
        except Exception as e:
            print(f"Error parsing MSG file {filepath}: {e}")
            raise

    def _extract_topic_keyword(self, subject, body):
        full_text = f"{subject}\n{body}"
        pattern1 = r'Topic\s*[:：]\s*([^\n\r,\.;]+)'
        pattern2 = r'Topic\s*[-–—]\s*([^\n\r,\.;]+)'
        pattern3 = r'[主主题題]题?\s*[:：]\s*([^\n\r,\.;]+)'
        for pattern in (pattern1, pattern2, pattern3):
            m = re.search(pattern, full_text, re.IGNORECASE)
            if m:
                keyword = m.group(1).strip()[:50]
                keyword = re.sub(r'\s+', ' ', keyword)
                if len(keyword) > 2:
                    return keyword
        return ''

    def _extract_body_eml(self, msg):
        body = ""
        if msg.is_multipart():
            for part in msg.walk():
                if part.get_content_type() == "text/plain":
                    try:
                        body += part.get_content()
                    except Exception:
                        pass
        else:
            try:
                body = msg.get_content()
            except Exception:
                body = str(msg)
        return body

    def _parse_conversation_thread(self, body_text):
        if not body_text:
            return {'count': 1, 'summary': '', 'dates': []}
        t = body_text.replace("\r\n", "\n").replace("\r", "\n")
        sections = re.split(r'(?i)(?:^|\n)(?:From:|De :|发件人:|보낸 사람:|-----Original Message-----)', t)
        entries, dates_collected = [], []
        hdr_from = re.compile(r'(?i)^From:\s*(.+)')
        hdr_date = re.compile(r'(?i)^(?:Date|Sent|Envoyé|보낸 날짜|날짜):\s*(.+)')
        for sec in sections:
            sec = sec.strip()
            if not sec:
                continue
            sender_line = ""; date_line = ""
            lines = sec.split("\n")
            for ln in lines[:20]:
                l = " ".join(ln.strip().split())
                if not l:
                    continue
                m_from = hdr_from.match(l)
                if m_from and not sender_line:
                    sender_line = m_from.group(1).strip()
                    continue
                m_date = hdr_date.match(l)
                if m_date and not date_line:
                    date_line = m_date.group(1).strip()
                    try:
                        parsed = self._parse_flexible_date(date_line)
                        if parsed:
                            dates_collected.append(parsed)
                            date_line = parsed.strftime("%A, %B %d, %Y %I:%M %p")
                    except Exception:
                        pass
            key_sentence = self._strip_greetings(sec)
            if key_sentence or sender_line or date_line:
                entries.append(self._format_sender_date_keyword(sender_line, date_line, key_sentence))
        if not entries:
            single = self._strip_greetings(t)
            if single:
                entries.append(self._format_sender_date_keyword("", "", single))
        summary_text = self._to_single_spaced_lines("\n".join(entries), max_lines=10, max_chars=1500)
        return {'count': max(1, len(entries)), 'summary': summary_text, 'dates': sorted(set(dates_collected))}

    def _parse_flexible_date(self, date_str):
        formats = [
            '%m/%d/%Y %I:%M:%S %p', '%m/%d/%Y %I:%M %p', '%d/%m/%Y %H:%M:%S',
            '%Y-%m-%d %H:%M:%S', '%a, %d %b %Y %H:%M:%S', '%B %d, %Y %I:%M %p',
        ]
        for fmt in formats:
            try:
                return datetime.strptime(date_str.strip(), fmt)
            except Exception:
                continue
        return None

    def _extract_conversation_summary(self, body_text, thread_count):
        # Deprecated but kept for compatibility
        if thread_count <= 1:
            lines = body_text.split('\n')
            meaningful = [ln.strip() for ln in lines if len(ln.strip()) > 20 and not ln.strip().startswith('>')]
            base = ' '.join(meaningful[:3])[:300]
            return self._to_single_spaced_lines(base, max_lines=3, max_chars=300)
        summary_parts = []
        sections = re.split(r'_{5,}|From:|Sent:|Original Message', body_text, flags=re.IGNORECASE)
        for section in sections[:5]:
            lines = section.split('\n')
            content_lines = []
            for line in lines:
                clean_line = " ".join(line.strip().split())
                if (
                    len(clean_line) > 30 and not clean_line.startswith('>') and
                    not clean_line.startswith('-----') and
                    not any(h in clean_line for h in ['Subject:', 'To:', 'Cc:', 'Date:', 'From:'])
                ):
                    content_lines.append(clean_line)
                    if len(content_lines) >= 2:
                        break
            if content_lines:
                summary_parts.extend(content_lines)
        full_summary = ' | '.join(summary_parts)
        return self._to_single_spaced_lines(full_summary or 'Conversation thread', max_lines=10, max_chars=500)

    def _extract_marker(self, text, pattern):
        m = re.search(pattern, text)
        return m.group(0) if m else ''

    # ---------------------- Classification ----------------------
    def classify_topics_with_bert(self):
        if self.use_fallback_clustering:
            print("Using improved K-Means (user selected fallback)")
            return self._classify_with_improved_kmeans()
        if not self.bert_model:
            if not self.initialize_bert():
                if self.use_fallback_clustering:
                    return self._classify_with_improved_kmeans()
                return False
        if len(self.emails) < 2:
            for e in self.emails:
                e['topic'] = 'Topic_1'; e['confidence'] = 1.0
            return True
        try:
            return self._classify_with_bert()
        except Exception as e:
            print(f"BERT processing failed: {e}")
            print("Falling back to improved K-Means...")
            return self._classify_with_improved_kmeans()

    def _classify_with_improved_kmeans(self):
        from sklearn.feature_extraction.text import TfidfVectorizer
        from sklearn.cluster import KMeans
        texts = []
        for e in self.emails:
            topic_hint = (" " + e.get('topic_keyword', '')) * 3 if e.get('topic_keyword') else ""
            body_clean = self._strip_greetings(e.get('body', '')[:1500])
            texts.append(f"{e['subject']} {e['subject']} {e['subject']}{topic_hint} {body_clean}")
        vectorizer = TfidfVectorizer(max_features=1000, stop_words='english', ngram_range=(1, 2), min_df=1, max_df=0.95)
        vectors = vectorizer.fit_transform(texts)
        n_clusters = min(5, max(2, len(self.emails) // 5))
        kmeans = KMeans(n_clusters=n_clusters, random_state=42, n_init=10)
        labels = kmeans.fit_predict(vectors)
        distances = kmeans.transform(vectors)
        for i, e in enumerate(self.emails):
            e['topic'] = f'Topic_{labels[i] + 1}'
            min_dist = distances[i][labels[i]]
            second_min_dist = np.partition(distances[i], 1)[1]
            e['confidence'] = round(1 - (min_dist / (min_dist + second_min_dist + 0.001)), 2)
        self._apply_topic_keywords()
        self._apply_correction_memory()
        self._apply_topicmap_to_emails()  # apply merges/renames for display
        print("Classification complete! (Improved K-Means)")
        return True

    def _classify_with_bert(self):
        print("Processing with BERT AI...")
        texts = []
        for e in self.emails:
            topic_hint = f" Topic: {e.get('topic_keyword', '')}" if e.get('topic_keyword') else ""
            clean_body = self._clean_for_embedding(e.get('body', '')[:1500])
            texts.append(f"{e['subject']} {e['subject']} {clean_body}{topic_hint}")
        print("Generating semantic embeddings (this may take a moment)...")
        try:
            from sklearn.preprocessing import normalize
            embeddings = self.bert_model.encode(texts, show_progress_bar=False, convert_to_numpy=True, batch_size=8)
            embeddings = normalize(embeddings)
        except Exception as e:
            print(f"BERT encoding failed: {e}")
            return self._classify_with_improved_kmeans()
        from sklearn.cluster import KMeans
        n_min, n_max = 2, min(10, max(3, len(self.emails)//2))
        candidates = list(range(n_min, n_max + 1))
        best_model = None
        try:
            from sklearn.metrics import silhouette_score
            best_score = -1
            for k in candidates:
                km = KMeans(n_clusters=k, random_state=42, n_init=10)
                labs = km.fit_predict(embeddings)
                score = silhouette_score(embeddings, labs, metric='euclidean')
                if score > best_score:
                    best_score, best_model = score, km
        except Exception:
            k = min(5, max(2, len(self.emails)//5))
            best_model = KMeans(n_clusters=k, random_state=42, n_init=10).fit(embeddings)
        labels = best_model.labels_
        from sklearn.preprocessing import normalize
        centroids = normalize(best_model.cluster_centers_)
        for i, e in enumerate(self.emails):
            e['topic'] = f'Topic_{labels[i] + 1}'
            e['confidence'] = round(float(np.dot(embeddings[i], centroids[labels[i]])), 2)
            e['embedding'] = embeddings[i]
        for e in self.emails:
            if e.get('confidence', 0) < self.min_confidence:
                e['topic'] = 'Uncertain'
        # Load learned centroids from per-Excel state (then fallback global)
        self._apply_learned_topics_to_embeddings(embeddings, min_conf=self.min_confidence)
        self._apply_topic_keywords()
        self._apply_correction_memory()
        self._apply_topicmap_to_emails()  # apply merges/renames for display
        print("Classification complete! (BERT)")
        return True

    def _apply_learned_topics_to_embeddings(self, embeddings, min_conf=0.65):
        import json
        state_paths = []
        if self.excel_path:
            state_paths.append(self._state_path_for_excel(self.excel_path))
        state_paths.append(Path('learning_state.json'))  # fallback for backward compatibility
        state = None
        for p in state_paths:
            try:
                if Path(p).exists():
                    with open(p, 'r', encoding='utf-8') as f:
                        state = json.load(f)
                        break
            except Exception as e:
                print(f"Could not read state from {p}: {e}")
        if not state or not state.get('topics'):
            return
        from sklearn.preprocessing import normalize
        topic_names = list(state['topics'].keys())
        C = np.vstack([np.array(state['topics'][t]) for t in topic_names])
        E = normalize(np.asarray(embeddings))
        sims = E @ C.T
        for i, e in enumerate(self.emails):
            best_idx = int(np.argmax(sims[i])); best_sim = float(sims[i, best_idx])
            if best_sim >= state.get('min_confidence', min_conf):
                e['topic'] = topic_names[best_idx]
                e['confidence'] = round(best_sim, 2)

    # ---------------------- Topic keywords & corrections ----------------------
    def _apply_topic_keywords(self):
        keyword_to_topic = {}
        for e in self.emails:
            kw = (e.get('topic_keyword', '') or '').lower().strip()
            if kw:
                if kw not in keyword_to_topic:
                    keyword_to_topic[kw] = e['topic']
                self.topic_keywords[kw] = e['topic']
        for e in self.emails:
            kw = (e.get('topic_keyword', '') or '').lower().strip()
            if kw and kw in keyword_to_topic:
                e['topic'] = keyword_to_topic[kw]
                e['confidence'] = 1.0

    def _apply_correction_memory(self):
        if not (self.correction_memory or self.correction_memory_by_subject):
            return
        for e in self.emails:
            mid = str(e.get('message_id', '')).strip()
            subj = (e.get('subject') or '').strip().lower()
            if mid and mid in self.correction_memory:
                corr = self.correction_memory[mid]
                e['topic'] = corr['corrected_topic']; e['confidence'] = 0.99
                continue
            if subj and subj in self.correction_memory_by_subject:
                corr = self.correction_memory_by_subject[subj]
                e['topic'] = corr['corrected_topic']; e['confidence'] = 0.98

    # ---------------------- TopicMap (Merge Into + Rename) ----------------------
    def _apply_topicmap_to_emails(self):
        if not self.topicmap_rules:
            return
        def resolve(topic: str) -> str:
            visited = set(); cur = topic
            while cur in self.topicmap_rules and self.topicmap_rules[cur].get('merge'):
                if cur in visited: break
                visited.add(cur)
                cur = self.topicmap_rules[cur]['merge']
            custom = self.topicmap_rules.get(cur, {}).get('name') or self.topicmap_rules.get(topic, {}).get('name')
            return custom or cur
        for e in self.emails:
            e['topic'] = resolve(e['topic'])

    # ---------------------- Excel I/O ----------------------
    def load_existing_excel(self, filepath):
        processed_ids = set()
        if not os.path.exists(filepath):
            return processed_ids
        try:
            wb = openpyxl.load_workbook(filepath)
            if 'Index' in wb.sheetnames:
                ws = wb['Index']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:
                        processed_ids.add(str(row[0]).strip())
            if 'Corrections' in wb.sheetnames:
                ws = wb['Corrections']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    mid = (row[0] or "").strip() if row[0] else ""
                    ai_topic = (row[1] or "").strip() if row[1] else ""
                    corrected = (row[2] or "").strip() if row[2] else ""
                    subj = (row[4] or "").strip().lower() if len(row) > 4 and row[4] else ""
                    if corrected:
                        if mid:
                            self.correction_memory[mid] = {'ai_topic': ai_topic, 'corrected_topic': corrected}
                        if subj:
                            self.correction_memory_by_subject[subj] = {'ai_topic': ai_topic, 'corrected_topic': corrected}
            # TopicMap rules
            self.topicmap_rules = {}
            if 'TopicMap' in wb.sheetnames:
                ws = wb['TopicMap']
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:
                        orig = str(row[0]).strip()
                        cname = (row[1] or '').strip()
                        merge = (row[2] or '').strip()
                        self.topicmap_rules[orig] = {'name': cname, 'merge': merge}
            wb.close()
        except Exception as e:
            print(f"Error loading existing Excel: {e}")
        return processed_ids

    def save_to_excel(self, filepath):
        if os.path.exists(filepath):
            wb = openpyxl.load_workbook(filepath)
        else:
            wb = openpyxl.Workbook(); wb.remove(wb.active)
        for name in ['Emails', 'Summary', 'TopicMap', 'Corrections', 'Index']:
            if name not in wb.sheetnames:
                wb.create_sheet(name)
        self._write_emails_sheet(wb['Emails'])
        self._write_summary_sheet(wb)
        self._write_topicmap_sheet(wb)
        self._write_corrections_sheet(wb['Corrections'])
        self._write_index_sheet(wb['Index'])
        wb.save(filepath); wb.close()

    def _write_emails_sheet(self, ws):
        if ws.max_row == 1 or ws.cell(1, 1).value != 'Date':
            headers = ['Date', 'Topic', 'Confidence', 'Topic Keyword', 'Subject', 'From',
                       'Thread Count', 'Conversation Summary', 'Issue', 'Milestone', 'Body Preview',
                       'Correct Topic', 'Message ID']
            ws.append(headers); self._style_header(ws)
        for e in self.emails:
            conv_summary = self._to_single_spaced_lines(e.get('thread_summary', '') or '', max_lines=10, max_chars=1500)
            body_preview = e.get('body_preview_line') or self._format_sender_date_keyword(
                e.get('from', ''), e.get('date').strftime("%A, %B %d, %Y %I:%M %p") if e.get('date') else '',
                self._strip_greetings(e.get('body','') or e.get('subject',''))
            )
            body_preview = self._to_single_spaced_lines(body_preview, max_lines=5, max_chars=1000)
            ws.append([
                e['date'].strftime('%Y-%m-%d %H:%M:%S'), e['topic'], e.get('confidence', 0.0),
                e.get('topic_keyword', ''), e['subject'], e['from'], e.get('thread_count', 1),
                conv_summary, e.get('issue', '') or '', e.get('milestone', '') or '', body_preview, '',
                str(e.get('message_id',''))
            ])
        self._auto_adjust_columns(ws)

    def _write_corrections_sheet(self, ws):
        if ws.max_row == 1 or ws.cell(1, 1).value != 'Message ID':
            ws.append(['Message ID', 'AI Topic', 'Corrected Topic', 'Date Corrected', 'Subject']); self._style_header(ws)
        self._auto_adjust_columns(ws)

    def _write_summary_sheet(self, wb):
        ws = wb['Summary']; ws.delete_rows(1, ws.max_row)
        ws.append(['Topic', 'Count', 'Avg Confidence', 'Latest Date', 'Topic Keywords', 'Mapped Topic Name'])
        self._style_header(ws)
        emails_ws = wb['Emails']; topic_data = {}
        for row in emails_ws.iter_rows(min_row=2, values_only=True):
            if not row[1]:
                continue
            topic = row[1]; date_str = row[0]; conf = row[2] or 0.0; tkw = row[3] or ''; thr = row[6] or 1
            if topic not in topic_data:
                topic_data[topic] = {'count': 0, 'latest': None, 'threads': 0, 'confidences': [], 'keywords': set()}
            topic_data[topic]['count'] += 1; topic_data[topic]['threads'] += thr; topic_data[topic]['confidences'].append(conf)
            if tkw: topic_data[topic]['keywords'].add(tkw)
            try:
                d = datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S')
                if not topic_data[topic]['latest'] or d > topic_data[topic]['latest']:
                    topic_data[topic]['latest'] = d
            except Exception:
                pass
        # For display only: map to Custom Name column
        topicmap = self._get_topic_mappings(wb)
        for topic in sorted(topic_data.keys()):
            latest = topic_data[topic]['latest']
            latest_str = latest.strftime('%Y-%m-%d %H:%M:%S') if latest else 'N/A'
            avg_conf = np.mean(topic_data[topic]['confidences']) if topic_data[topic]['confidences'] else 0
            keywords = ', '.join(sorted(topic_data[topic]['keywords']))
            mapped_name = topicmap.get(topic, '')
            ws.append([topic, topic_data[topic]['count'], round(avg_conf, 2), latest_str, keywords, mapped_name])
        self._auto_adjust_columns(ws)

    def _write_topicmap_sheet(self, wb):
        emails_ws = wb['Emails']; all_topics = set()
        for row in emails_ws.iter_rows(min_row=2, values_only=True):
            if row[1]: all_topics.add(row[1])
        ws = wb['TopicMap']; existing = {}
        if ws.max_row > 1 and ws.cell(1, 1).value == 'Original Topic':
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]: existing[row[0]] = {'custom_name': row[1] or '', 'merge_into': row[2] or ''}
        ws.delete_rows(1, ws.max_row)
        ws.append(['Original Topic', 'Custom Name', 'Merge Into']); self._style_header(ws)
        for t in sorted(all_topics):
            m = existing.get(t, {'custom_name': '', 'merge_into': ''})
            ws.append([t, m['custom_name'], m['merge_into']])
        self._auto_adjust_columns(ws)

    def _write_index_sheet(self, ws):
        if ws.max_row == 1 or ws.cell(1, 1).value != 'Message ID':
            ws.append(['Message ID', 'Processed Date', 'Filename']); self._style_header(ws)
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        for e in self.emails:
            ws.append([str(e['message_id']).strip(), now, Path(e['filepath']).name])
        self._auto_adjust_columns(ws)

    def _get_topic_mappings(self, wb):
        mappings = {}
        if 'TopicMap' in wb.sheetnames:
            ws = wb['TopicMap']
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]: mappings[row[0]] = row[1] or ''
        return mappings

    def _auto_adjust_columns(self, ws):
        for column_cells in ws.columns:
            length = 0; column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if cell.value:
                        length = max(length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[column].width = min(max(length + 2, 10), 80)

    def _style_header(self, ws):
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        for cell in ws[1]:
            cell.fill = header_fill; cell.font = header_font; cell.alignment = Alignment(horizontal='center')

    # ---------------------- Learning from Excel (per-Excel state + merges) ----------------------
    def learn_from_excel_corrections(self, filepath):
        if not os.path.exists(filepath):
            return 0
        try:
            wb = openpyxl.load_workbook(filepath)
            ws_emails = wb['Emails']; ws_corr = wb['Corrections']
            headers = {ws_emails.cell(1, c).value: c for c in range(1, ws_emails.max_column + 1)}
            subj_col = headers.get('Subject'); body_col = headers.get('Body Preview')
            correct_col = headers.get('Correct Topic'); msgid_col = headers.get('Message ID')
            if not correct_col:
                wb.close(); return 0
            synced = 0
            for r in range(2, ws_emails.max_row + 1):
                corrected = ws_emails.cell(r, correct_col).value
                if corrected and str(corrected).strip():
                    ai_topic = ws_emails.cell(r, headers.get('Topic')).value if headers.get('Topic') else ''
                    subj = ws_emails.cell(r, subj_col).value if subj_col else ''
                    msg_id = ws_emails.cell(r, msgid_col).value if msgid_col else f"row_{r}"
                    ws_corr.append([str(msg_id).strip(), ai_topic, str(corrected).strip(),
                                    datetime.now().strftime('%Y-%m-%d %H:%M:%S'), subj])
                    synced += 1
            if synced > 0:
                wb.save(filepath)
            # Build centroids from Corrections
            ws_corr = wb['Corrections']; corrected_rows = []
            subj_idx = headers.get('Subject', 5); body_idx = headers.get('Body Preview', 11)
            emails_rows = list(ws_emails.iter_rows(min_row=2, values_only=True))
            for row in ws_corr.iter_rows(min_row=2, values_only=True):
                ctopic = (row[2] or '').strip(); subj = (row[4] or '').strip() if len(row) > 4 and row[4] else ''
                if not ctopic or not subj:
                    continue
                for rr in emails_rows:
                    e_subj = (rr[subj_idx - 1] or '').strip()
                    if e_subj == subj:
                        e_body = rr[body_idx - 1] or ''
                        corrected_rows.append((ctopic, subj, e_body))
                        break
            # TopicMap rules (for merging centroids)
            self.topicmap_rules = {}
            if 'TopicMap' in wb.sheetnames:
                ws_tm = wb['TopicMap']
                for row in ws_tm.iter_rows(min_row=2, values_only=True):
                    if row[0]:
                        self.topicmap_rules[str(row[0]).strip()] = {
                            'name': (row[1] or '').strip(), 'merge': (row[2] or '').strip()
                        }
            wb.close()
            if not corrected_rows:
                return synced
            self._ensure_bert()
            from sklearn.preprocessing import normalize
            texts, labels = [], []
            for ctopic, subj, body in corrected_rows:
                texts.append(f"{subj}\n\n{self._clean_for_embedding(body)}"); labels.append(ctopic)
            embs = self.bert_model.encode(texts, convert_to_numpy=True, show_progress_bar=False, batch_size=8)
            embs = normalize(embs)
            from collections import defaultdict
            sums = defaultdict(lambda: None); counts = defaultdict(int)
            for vec, lab in zip(embs, labels):
                if sums[lab] is None: sums[lab] = vec.copy()
                else: sums[lab] += vec
                counts[lab] += 1
            centroids = {}
            for lab in sums:
                c = sums[lab] / max(1, counts[lab]); c = c / (np.linalg.norm(c) + 1e-9)
                centroids[lab] = c
            # Apply TopicMap merges to centroids
            merged = {}; group = {}
            def target_of(t: str) -> str:
                cur = t; seen = set()
                while cur in self.topicmap_rules and self.topicmap_rules[cur].get('merge'):
                    if cur in seen: break
                    seen.add(cur)
                    cur = self.topicmap_rules[cur]['merge']
                return cur
            for name, vec in centroids.items():
                tgt = target_of(name)
                group.setdefault(tgt, []).append(vec)
            for tgt, vecs in group.items():
                m = np.mean(np.vstack(vecs), axis=0)
                m = m / (np.linalg.norm(m) + 1e-9)
                disp = self.topicmap_rules.get(tgt, {}).get('name') or tgt
                merged[disp] = m.tolist()
            # Save per-Excel learning state
            state_path = self._state_path_for_excel(filepath)
            self.state_path = state_path
            state = {"topics": merged, "min_confidence": self.min_confidence,
                     "version": datetime.now().strftime("%Y-%m-%dT%H:%M:%S")}
            import json
            with open(state_path, 'w', encoding='utf-8') as f:
                json.dump(state, f, ensure_ascii=False, indent=2)
            print(f"Saved learning state to {state_path}")
            return synced
        except Exception as e:
            print(f"Error learning from corrections: {e}")
            return 0

    def _clean_for_embedding(self, text: str) -> str:
        if not isinstance(text, str): return ""
        t = text.replace("\r", "\n")
        bad_prefixes = ("From:", "Sent:", "De :", "Envoyé", "주의", "CAUTION", "Sujet:", "Subject:")
        lines = []
        for ln in t.split("\n"):
            l = ln.strip()
            if not l or any(l.startswith(bp) for bp in bad_prefixes) or l.startswith(">"):
                continue
            lines.append(l)
            if len(lines) >= 300:
                break
        return " ".join(lines)


# ---------------------- GUI ----------------------
class EmailOrganizerGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Email Topic Organizer - BERT AI with Learning (Patched v2)")
        self.root.geometry("800x650")
        self.organizer = EmailOrganizer()
        self.dropped_files = []
        self.drag_drop_available = False
        self.setup_ui(); self.setup_drag_drop()

    def setup_ui(self):
        title = tk.Label(self.root, text="📧 Email Topic Organizer - Smart AI", font=('Arial', 16, 'bold'))
        title.pack(pady=10)
        info = tk.Label(self.root,
                        text=("✨ BERT AI • Cosine Confidence • Learns from Excel • Multi-language\n"
                              "💡 Tip: Add 'Topic: YourCategory' in email body to force grouping"),
                        justify=tk.CENTER, fg='#666')
        info.pack(pady=5)
        drop_frame = tk.Frame(self.root, bg='#e8f4f8', relief=tk.SOLID, borderwidth=2)
        drop_frame.pack(padx=20, pady=10, fill=tk.BOTH, expand=True)
        drop_text = "📁 Select .eml or .msg files using Browse button\n\nSupported: .eml, .msg | Thread detection: Auto"
        if self.drag_drop_available:
            drop_text = "📁 Drop email files here or use Browse button\n\nSupported: .eml, .msg | Thread detection: Auto"
        self.drop_label = tk.Label(drop_frame, text=drop_text, bg='#e8f4f8', font=('Arial', 12), justify=tk.CENTER)
        self.drop_label.pack(expand=True)
        list_frame = tk.Frame(self.root); list_frame.pack(padx=20, pady=5, fill=tk.BOTH, expand=True)
        scrollbar = tk.Scrollbar(list_frame); scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.file_listbox = tk.Listbox(list_frame, yscrollcommand=scrollbar.set, height=8)
        self.file_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True); scrollbar.config(command=self.file_listbox.yview)
        btn_frame = tk.Frame(self.root); btn_frame.pack(pady=10)
        self.browse_btn = tk.Button(btn_frame, text="📂 Browse Files", command=self.browse_files, width=15)
        self.browse_btn.pack(side=tk.LEFT, padx=5)
        self.clear_btn = tk.Button(btn_frame, text="🗑️ Clear List", command=self.clear_files, width=15)
        self.clear_btn.pack(side=tk.LEFT, padx=5)
        self.process_btn = tk.Button(btn_frame, text="⚡ Process & Export", command=self.process_emails,
                                     width=15, bg='#4CAF50', fg='white', font=('Arial', 10, 'bold'))
        self.process_btn.pack(side=tk.LEFT, padx=5)
        btn_frame2 = tk.Frame(self.root); btn_frame2.pack(pady=5)
        self.learn_btn = tk.Button(btn_frame2, text="🧠 Learn from Excel", command=self.learn_from_corrections,
                                   width=18, bg='#2196F3', fg='white')
        self.learn_btn.pack(side=tk.LEFT, padx=5)
        self.help_btn = tk.Button(btn_frame2, text="❓ How to Use", command=self.show_help, width=18)
        self.help_btn.pack(side=tk.LEFT, padx=5)
        self.status_label = tk.Label(self.root, text="Ready - BERT AI will download on first use (~420MB)", fg='green')
        self.status_label.pack(pady=5)
        self.progress = ttk.Progressbar(self.root, mode='indeterminate'); self.progress.pack(padx=20, pady=5, fill=tk.X)

    def setup_drag_drop(self):
        try:
            from tkinterdnd2 import TkinterDnD
            self.root.drop_target_register('DND_Files')
            self.root.dnd_bind('<<Drop>>', self.on_drop)
            self.drag_drop_available = True
            self.drop_label.config(text="📁 Drop email files here or use Browse button\n\nSupported: .eml, .msg | Thread detection: Auto")
            self.status_label.config(text="Ready - Drag & drop enabled | BERT AI ready")
        except Exception as e:
            print(f"Drag and drop not available: {e}")
            self.drag_drop_available = False

    def on_drop(self, event):
        try:
            files = self.root.tk.splitlist(event.data)
            self.add_files(files)
        except Exception as e:
            print(f"Error handling drop: {e}")

    def browse_files(self):
        files = filedialog.askopenfilenames(title="Select Email Files",
                                            filetypes=[("Email files", "*.eml *.msg"), ("All files", "*.*")])
        if files:
            self.add_files(files)

    def add_files(self, files):
        for filepath in files:
            if filepath.lower().endswith(('.eml', '.msg')) and filepath not in self.dropped_files:
                self.dropped_files.append(filepath)
                self.file_listbox.insert(tk.END, Path(filepath).name)
        self.update_status(f"{len(self.dropped_files)} files ready")

    def clear_files(self):
        self.dropped_files = []
        self.file_listbox.delete(0, tk.END)
        self.update_status("Ready")

    def process_emails(self):
        if not self.dropped_files:
            messagebox.showwarning("No Files", "Please add email files first"); return
        excel_path = filedialog.asksaveasfilename(title="Save Excel File As", defaultextension=".xlsx",
                                                  filetypes=[("Excel files", "*.xlsx")])
        if not excel_path:
            return
        # Remember Excel path (for per-Excel learning state)
        self.organizer.excel_path = excel_path
        self.organizer.state_path = self.organizer._state_path_for_excel(excel_path)
        self.process_btn.config(state=tk.DISABLED); self.browse_btn.config(state=tk.DISABLED); self.clear_btn.config(state=tk.DISABLED)

        def process_thread():
            try:
                processed_ids = self.organizer.load_existing_excel(excel_path)
                new_emails = 0; skipped = 0; failed_files = []; total_threads = 0
                self.root.after(0, lambda: self.update_status(f"Parsing {len(self.dropped_files)} email files..."))
                for filepath in self.dropped_files:
                    try:
                        data = self.organizer.parse_email_file(filepath)
                        if data:
                            mid = str(data['message_id']).strip()
                            if mid not in processed_ids:
                                self.organizer.emails.append(data); new_emails += 1
                                total_threads += data.get('thread_count', 1)
                            else:
                                skipped += 1
                    except Exception as e:
                        failed_files.append((Path(filepath).name, str(e)))
                        print(f"Failed to parse {filepath}: {e}")
                if new_emails == 0:
                    self.root.after(0, lambda: self._show_no_new_emails(skipped)); return
                self.root.after(0, lambda: self.update_status(f"🧠 BERT processing {new_emails} emails — this may take a moment..."))
                self.root.after(0, lambda: self.progress.start(10))
                success = self.organizer.classify_topics_with_bert()
                if not success:
                    self.root.after(0, lambda: self._show_error("BERT classification failed")); return
                self.root.after(0, lambda: self.update_status("💾 Saving to Excel..."))
                self.organizer.save_to_excel(excel_path)
                self.root.after(0, lambda: self._show_success(new_emails, total_threads, skipped, failed_files, excel_path))
                self.root.after(0, lambda: self.clear_files()); self.organizer.emails = []
            except Exception as e:
                import traceback
                print("Error processing emails:\n" + traceback.format_exc())
                self.root.after(0, lambda: self._show_error(str(e)))
            finally:
                self.root.after(0, lambda: self.process_btn.config(state=tk.NORMAL))
                self.root.after(0, lambda: self.browse_btn.config(state=tk.NORMAL))
                self.root.after(0, lambda: self.clear_btn.config(state=tk.NORMAL))
                self.root.after(0, lambda: self.progress.stop())

        import threading
        thread = threading.Thread(target=process_thread, daemon=True); thread.start()
        self.update_status("⚙️ Processing started in background..."); self.progress.start()

    def _show_no_new_emails(self, skipped):
        self.progress.stop(); info_msg = "All selected emails have already been processed or failed to parse."
        if skipped > 0: info_msg += f"\n\nSkipped: {skipped} duplicates"
        messagebox.showinfo("No New Emails", info_msg); self.update_status("No new emails to process")

    def _show_error(self, error_msg):
        self.progress.stop(); messagebox.showerror("Error", f"Failed to process emails:\n{error_msg}")
        self.update_status("❌ Error occurred")

    def _show_success(self, new_emails, total_threads, skipped, failed_files, excel_path):
        self.progress.stop()
        info_msg = f"✅ Processed: {new_emails} emails ({total_threads} messages in threads)\n"
        info_msg += f"🎯 Topic keywords found: {len(self.organizer.topic_keywords)}\n"
        if skipped > 0: info_msg += f"⏭️ Skipped: {skipped} duplicates\n"
        if failed_files: info_msg += f"❌ Failed: {len(failed_files)} files\n"
        info_msg += f"\n📁 Saved to:\n{excel_path}\n\n"
        info_msg += "💡 Tip: Fill 'Correct Topic' in Excel and click 'Learn from Excel' to improve AI!"
        messagebox.showinfo("Success!", info_msg)
        self.update_status(f"✅ Success! Processed {new_emails} emails with BERT AI")

    def learn_from_corrections(self):
        excel_path = filedialog.askopenfilename(title="Select Excel File to Learn From", filetypes=[("Excel files", "*.xlsx")])
        if not excel_path:
            return
        self.organizer.excel_path = excel_path
        self.organizer.state_path = self.organizer._state_path_for_excel(excel_path)
        self.update_status("Learning from your corrections..."); self.progress.start(); self.root.update()
        try:
            corrections_found = self.organizer.learn_from_excel_corrections(excel_path)
            self.progress.stop()
            if corrections_found > 0:
                messagebox.showinfo(
                    "Learning Complete",
                    "Learned from your corrections! A per-file learning state was saved next to your Excel."
                )
                self.update_status(f"✅ Learned from {corrections_found} corrections")
            else:
                messagebox.showinfo(
                    "No Corrections Found",
                    "No corrections found in 'Correct Topic'.\nOpen Excel → fill 'Correct Topic' → Save → Learn again."
                )
                self.update_status("No corrections to learn from")
        except Exception as e:
            self.progress.stop(); messagebox.showerror("Error", f"Failed to learn from Excel:\n{str(e)}")
            self.update_status("❌ Error occurred")

    def show_help(self):
        help_text = (
            """
📧 EMAIL TOPIC ORGANIZER - SMART AI HELP (Patched v2)

🎯 HOW IT WORKS:
1. Select .eml or .msg files
2. Click "Process & Export"
3. AI groups emails by topic (cosine confidence, auto-K)
4. Export to Excel with 5 sheets

🧠 LEARNING (per Excel file):
• Fill "Correct Topic" in Emails sheet
• Click "Learn from Excel" → saves <YourWorkbook>.learning_state.json next to your Excel
• Next runs load this state automatically

🧩 TOPICMAP (Merge Into + Rename):
• Use TopicMap sheet to merge/rename topics (e.g., Topic_3 → tool PO)
• Merges affect learned centroids and final topics

📝 FORMATTING:
• Body Preview & Conversation Summary: "sender | date | key sentence"
• Greetings like "안녕하세요" / "Hello" removed

🔧 NOTES:
• Confidence < 0.65 → Uncertain (review in Excel)
• 100% local & private
            """
        )
        win = tk.Toplevel(self.root); win.title("Help - How to Use"); win.geometry("600x700")
        txt = tk.Text(win, wrap=tk.WORD, padx=10, pady=10); txt.pack(fill=tk.BOTH, expand=True)
        txt.insert(1.0, help_text); txt.config(state=tk.DISABLED)
        tk.Button(win, text="Close", command=win.destroy).pack(pady=10)

    def update_status(self, message):
        self.status_label.config(text=message); self.root.update()


def main():
    try:
        from tkinterdnd2 import TkinterDnD
        root = TkinterDnD.Tk()
    except Exception:
        root = tk.Tk()
    app = EmailOrganizerGUI(root)
    root.mainloop()


if __name__ == '__main__':
    main()
